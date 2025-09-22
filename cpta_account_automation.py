import win32com.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date, datetime
import os
from sqlalchemy import create_engine
import pandas as pd


#CPTA CRM昨日9时至今日9时数据汇总Excel
output_path = 'C:\\Users\\Frank W\\OneDrive - Logtec Innovation Limited\\Desktop\\Daily_Data\\data.xlsx'
print(datetime.today())

# Replace with your actual credentials
db_user = "new_frank.w"
db_pass = "XH5mJ!MEHi#QFJN3"
db_host = "proxy.cptmarkets.com"       # e.g., "localhost" or "192.168.1.100"
db_port = "50713"
db_name = "cpt"
# SQL for query database
query_cpta_top10_lots = """
SELECT
  login,
  SUM(std_lots) AS total_lots,
  business_group,
  SUM(profit_usd) AS total_profit_usd
FROM cpt.trade_transaction2509
WHERE std_lots BETWEEN 0.1 AND 50
  AND ticket_time BETWEEN 0 AND 180
  AND close_time BETWEEN '2025-09-21 09:00:00' AND '2025-09-22 09:00:00'
GROUP BY login, business_group
ORDER BY total_lots DESC
LIMIT 10;
"""
query_cpta_top10_profit = """
SELECT
  login,
  SUM(profit_usd) AS total_profit_usd,
  business_group,
  SUM(std_lots) AS total_lots
FROM cpt.trade_transaction2509
WHERE std_lots BETWEEN 0.1 AND 50
  AND ticket_time BETWEEN 0 AND 180
  AND close_time BETWEEN '2025-09-21 09:00:00' AND '2025-09-22 09:00:00'
GROUP BY login, business_group
ORDER BY total_profit_usd DESC
LIMIT 10;
"""

# SQL for EA trades
query_EA_trades = """
SELECT login, open_time, cmd, symbol, lots, open_price, close_time, close_price, business_group, profit_usd AS profit, comment 
    FROM cpt.trade_transaction2509 
    WHERE close_time >= '2025-09-21 09:00:00' AND close_time <= '2025-09-22 09:00:00' AND comment LIKE '期迹%';
"""



def load_outlook_email_attachment(subject_keyword=None, sender_email=None, received_date=None, file_type=".csv",
                            save_folder="C:\\Users\\Frank W\\OneDrive - Logtec Innovation Limited\\Desktop\\Data Integration\\Abnormal Accounts"):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # 6 = Inbox
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  # Sort by most recent

    for message in messages:
        if subject_keyword and subject_keyword not in message.Subject:
            continue
        if sender_email and sender_email.lower() != message.SenderEmailAddress.lower():
            continue
        if received_date and message.ReceivedTime.date() != received_date:
            continue

        for attachment in message.Attachments:
            if attachment.FileName.endswith(file_type):
                filepath = os.path.join(save_folder, attachment.FileName)
                attachment.SaveAsFile(filepath)
                print(f"Saved: {filepath}")

                # Load into DataFrame
                if file_type == ".csv":
                    return pd.read_csv(filepath)
                elif file_type == ".xlsx":
                    return pd.read_excel(filepath)

        break  # Stop after first matching email

    print("No matching email or attachment found.")
    return None


def calculate_centroid_data(df):
    if df is not None:
        '''Top 20 Volume-Lots Centroid accounts'''
        top_tops_columns = ["SID", "Login", "Group", "Volume - Lots", "Net Profit"]
        df_lots_filtered = df[top_tops_columns]
        df_cpta_lots_filtered = df_lots_filtered[~df_lots_filtered["SID"].isin(["mt4_3", "mt5_1"])]
        df_top20_lots = df_cpta_lots_filtered.sort_values(by="Volume - Lots", ascending=False).head(20)
        # save_dataframe_to_excel(df_top20_lots)
        # print(df_top10_lots)

        '''Top 20 Net Profit Centroid accounts'''
        top_profit_columns = ["SID", "Login", "Group", "Net Profit", "Volume - Lots"]
        df_profit_filtered = df[top_profit_columns]
        df_cpta_profit_filtered = df_profit_filtered[~df_profit_filtered["SID"].isin(["mt4_3", "mt5_1"])]
        df_top20_profit = df_cpta_profit_filtered.sort_values(by="Net Profit", ascending=False).head(20)
        # append_dataframe_to_excel(df_top20_profit)
        return df_top20_lots, df_top20_profit



# def write_dataframe_to_excel(df_top20_lots, df_top20_profit):
#     df_top20_lots.to_excel(output_path, index=False, startrow=1)  #create new excel file
#     # Step 1: Write both DataFrames to the same sheet
#     with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#         df_top20_lots.to_excel(writer, sheet_name="Sheet1", startrow=1, index=False)
#         last_row = len(df_top20_lots) + 5
#         df_top20_profit.to_excel(writer, sheet_name="Sheet1", startrow=last_row, index=False)
#
#     # Step 2: Reopen workbook to style it
#     wb = load_workbook(output_path)
#     ws = wb["Sheet1"]
#
#     # Header for first DataFrame
#     ws.merge_cells("A1:E1")
#     ws["A1"].value = "Centroid昨日交易量前20账户"
#     ws["A1"].font = Font(bold=True, size=16)
#     ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
#
#     # Header for second DataFrame
#     ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
#     ws.cell(row=last_row, column=1).value = "Centroid昨日盈利前20账户"
#     ws.cell(row=last_row, column=1).font = Font(bold=True, size=16)
#     ws.cell(row=last_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
#
#     # Save final styled workbook
#     wb.save(output_path)
def write_dataframe_to_excel(df_top20_lots, df_top20_profit):
    # Save the first DataFrame normally
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
        df_top20_lots.to_excel(writer, sheet_name="Sheet1", startrow=1, index=False)

    # Convert second DataFrame to rows
    rows = list(dataframe_to_rows(df_top20_profit, index=False, header=True))

    # Starting column for second DataFrame (e.g., column F = 6)
    start_col = 6 + 2
    start_row = 3  # Same row as first DataFrame's data

    # Write second DataFrame manually
    for r_idx, row in enumerate(rows, start=start_row - 1):  # -1 to include header
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Header for first table (A1 to E1)
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Centroid昨日交易量前20账户"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Header for second table (H1 to L1)
    ws.merge_cells("H1:L1")
    ws["H1"].value = "Centroid昨日盈利前20账户"
    ws["H1"].font = Font(bold=True, size=16)
    ws["H1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["H1"].alignment = Alignment(horizontal="center", vertical="center")
    # Save workbook
    wb.save(output_path)

    # Return the last row
    last_row = len(df_top20_lots) + 5
    return last_row


# Query database to get CPTA CRM data
def query_cpta_crm_data():
    # Create SQLAlchemy engine
    engine = create_engine(f"mysql+pymysql://{db_user}:{db_pass}@{db_host}:{db_port}/{db_name}")
    df_top10_lots = pd.read_sql(query_cpta_top10_lots, engine)
    df_top10_profit = pd.read_sql(query_cpta_top10_profit, engine)
    return df_top10_lots, df_top10_profit



def write_cpta_data_to_excel(last_row, df_top10_lots, df_top10_profit):
    start_col_lots = 1  # Column A
    start_col_profit = 8  # Column H
    # Convert DataFrames to Rows
    rows_lots = list(dataframe_to_rows(df_top10_lots, index=False, header=True))
    rows_profit = list(dataframe_to_rows(df_top10_profit, index=False, header=True))
    # Write df_cpta_lots
    for r_idx, row in enumerate(rows_lots, start=last_row):
        for c_idx, value in enumerate(row, start=start_col_lots):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Write df_cpta_profit
    for r_idx, row in enumerate(rows_profit, start=last_row):
        for c_idx, value in enumerate(row, start=start_col_profit):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Header for df_cpta_lots
    ws.merge_cells(start_row=last_row - 1, start_column=start_col_lots,
                   end_row=last_row - 1, end_column=start_col_lots + len(df_cpta_lots.columns) - 1)
    ws.cell(row=last_row - 1, column=start_col_lots).value = "180秒CPTA CRM交易量前10账户"
    ws.cell(row=last_row - 1, column=start_col_lots).font = Font(bold=True, size=14)
    ws.cell(row=last_row - 1, column=start_col_lots).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=last_row - 1, column=start_col_lots).alignment = Alignment(horizontal="center", vertical="center")

    # Header for df_cpta_profit
    ws.merge_cells(start_row=last_row - 1, start_column=start_col_profit,
                   end_row=last_row - 1, end_column=start_col_profit + len(df_cpta_profit.columns) - 1)
    ws.cell(row=last_row - 1, column=start_col_profit).value = "180秒CPTA CRM盈利前10账户"
    ws.cell(row=last_row - 1, column=start_col_profit).font = Font(bold=True, size=14)
    ws.cell(row=last_row - 1, column=start_col_profit).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=last_row - 1, column=start_col_profit).alignment = Alignment(horizontal="center", vertical="center")

    # Save Workbook
    wb.save(output_path)
    last_row = last_row + len(df_top10_lots) + 5
    return last_row



def save_dataframe_to_excel(df: pd.DataFrame) -> pd.DataFrame:
    # Save the DataFrame to Excel
    df.to_excel(output_path, index=False, startrow=1)  # Leave space for header
    # Load workbook and access sheet
    wb = load_workbook(output_path)
    ws = wb.active
    # Add custom header above the data
    ws.merge_cells("A1:E1")  # Merge cells A1 to E1
    ws["A1"] = "Centroid昨日交易量前20账户"  # Set the header text
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # Save changes
    wb.save(output_path)


def append_dataframe_to_excel(df: pd.DataFrame) -> pd.DataFrame:
    wb = load_workbook(output_path)
    ws = wb.active
    last_row = ws.max_row + 2  # Add spacing between tables
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # writer.book = wb
        # writer.sheets = {ws.title: ws}
        df.to_excel(writer, sheet_name=ws.title, startrow=last_row, index=False)

    print("Writing to row:", last_row - 1)
    ws.cell(row=last_row - 1, column=1).value = "Second DataFrame Summary"
    ws.cell(row=last_row - 1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=last_row - 1, column=1).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    ws.cell(row=last_row - 1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=last_row - 1, start_column=1, end_row=last_row - 1, end_column=len(df.columns))
    # ws["A24"] = "Centroid昨日盈利前20账户"  # Set the header text
    # wb.save(output_path)


def open_excel_and_write(df: pd.DataFrame) -> pd.DataFrame:
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        writer.book = load_workbook
        writer.sheets = {ws.title: ws for ws in load_workbook.worksheets}

        # Write the second DataFrame to a new sheet
        df.to_excel(writer, sheet_name="Summary", index=False)


if __name__ == "__main__":
    index = 0
    # Load workbook and sheet
    wb = load_workbook(output_path)
    ws = wb["Sheet1"]

    df = load_outlook_email_attachment(
        subject_keyword="All-client",
        sender_email="rms-noreply=centroidsol.com@mg.centroidsol.com",
        received_date=datetime(2025, 9, 22).date(),
        file_type=".xlsx"
    )

    # Centroid前20账户
    df_lots, df_profit = calculate_centroid_data(df)
    index = write_dataframe_to_excel(df_lots, df_profit)

    # CPTA CRM前10账户
    df_cpta_lots, df_cpta_profit = query_cpta_crm_data()
    index = write_cpta_data_to_excel(index, df_cpta_lots, df_cpta_profit)

